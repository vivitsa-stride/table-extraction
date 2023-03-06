import tabula
import PyPDF2
import os
import csv
import openpyxl
import pandas as pd
import shutil
from pathlib import Path
import numpy as np
from concurrent.futures import ProcessPoolExecutor
from itertools import repeat


import tabula
import PyPDF2
import os
from openpyxl import Workbook
import csv
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import shutil
from pathlib import Path
import numpy as np


class Extract():


    def __init__(self, inputFilePath, pageNumber, bbox = None, outputDirectory = None ) -> None:
        self.inputFilePath = inputFilePath
        self.pageNumber = pageNumber
        self.bbox = bbox
        
        if outputDirectory == None:
            self.outputDirectory = self._create_output_directory()
        else:
            self.outputDirectory = outputDirectory
        self.results = self.extract_tables()


    def _create_output_directory(self):

        filename = self.inputFilePath
        outputfolder = filename.split('/')[-1]
        dir ="/home/vaibhav/Desktop/tables_repo/resDocs/"
        outputfolder = outputfolder[0:-4]
        output_dir = dir+outputfolder
        if os.path.exists(output_dir):
            pass
        else:          
            os.mkdir(output_dir)
        return output_dir
    
    def extract_tables(self):
        # all tables from all pages from given file
        if self.pageNumber:
            # try:
                if self.bbox:
                    resDir = self.extract_bbox(self.inputFilePath, self.pageNumber, self.bbox)
                else:
                    resDir = self.extract_page(self.inputFilePath, self.pageNumber)

            # except: 
            #     pass

        return resDir


    def extract_bbox(self,filename,page,bbox):
        
        output_file_name = self.outputDirectory +"/" +"within_bbox_page_"+ str(page)
        extension = ".xlsx"
        if os.path.exists(output_file_name+extension):
            output_file_name = output_file_name+"(1)"
        else:
            pass
        tabula.convert_into(filename, output_file_name, area = bbox, output_format="csv", pages=page,lattice=True)
        rows=[]
        with open(output_file_name) as csvfile:  
            csvreader = csv.reader(csvfile,delimiter=',')
            j = 0
            for row in csvreader:
                j += 1
                i=0
                if(len(row)<=1):
                    continue
                while i < len(row):
                    if 'Unnamed' in row[i]:
                        row[i] = "  "
                    i += 1
                rows.append(row)  
        with open(output_file_name, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=',')
            writer.writerows(rows)
        dest_wb = Workbook()
        # Create new sheet in destination Workbook
        dest_wb.create_sheet(output_file_name.split('/')[-1])
        dest_ws = dest_wb[output_file_name.split('/')[-1]]


        # Read source data
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(output_file_name) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        wb.save(output_file_name + ".xlsx")
        os.remove(output_file_name)
        return output_file_name

    def extract_page(self,filename,page):

        # try:
        df_str = tabula.read_pdf(filename,pages=page,stream=True)
            

        for j in range(len(df_str)):
                name = self.outputDirectory +"/" + "all_in_page_"+str(page) + "-" +str(j)
                extension = ".xlsx"
                output_file_name = name+extension
                if os.path.exists(output_file_name):
                    output_file_name = name+"(1)"+extension
                df_str[j].to_csv(output_file_name)

        # except Exception as e:
        #     print(e)
        return output_file_name



    def convert_to_excel(self,file):
            rows = []
            with open(file) as csvfile:
                csvreader = csv.reader(csvfile,delimiter=',')
                j = 0
                for row in csvreader:
                    j += 1
                    i=0
                    while i < len(row):
                        if 'Unnamed' in row[i]:
                            row[i] = "  "
                        i += 1
                    rows.append(row)
            with open(file, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile, delimiter=',')
                writer.writerows(rows)

            # Read source data
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(file) as f:
                reader = csv.reader(f, delimiter=',')
                for row in reader:
                    ws.append(row)
            wb.save(file + ".xlsx")
            os.remove(file)

from time import time
start_time = time()
e = Extract(inputFilePath="/home/vaibhav/Desktop/tables_repo/Moodys Sfg2/batch1/9193_20190620.pdf", pageNumber = 4, bbox = None, outputDirectory=None)
print("results are in: ", e.results)

end_time = time()
print("time taken = ",end_time-start_time)







    # def extract_all_pages(self,filename):
    #     outputfile = filename.split('/')[-1]
    #     dir = filename.replace(outputfile,"")
    #     outputfile = outputfile[0:-4]
    #     dir_containing_files = dir + outputfile
    #     if os.path.exists(dir_containing_files):
    #         pass
    #     else:
    #         os.mkdir(dir_containing_files)
    #     file = open(filename, 'rb')
    #     readpdf = PyPDF2.PdfReader(file)
    #     pages = len(readpdf.pages)
    #     i = range(1,pages+1)
    #     with ProcessPoolExecutor(max_workers=min(os.cpu_count(),pages)) as exe:
    #         exe.map(self.extract_table_from_page,repeat(filename),i)
    #         exe.shutdown()

    #     files = os.listdir(dir_containing_files)
    #     files = [os.path.join(dir_containing_files, f) for f in files]
    #     files.sort(key=os.path.getctime)

    #     with ProcessPoolExecutor(max_workers=min(os.cpu_count(),len(files))) as exe:
    #         exe.map(self.convert_to_excel,files)
    #         exe.shutdown()


    #     return dir_containing_files
