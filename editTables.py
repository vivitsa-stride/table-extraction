import pandas as pd
import openpyxl
import openpyxl
from openpyxl.utils.cell import get_column_letter


def change_cell(file, cell_val):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for dict in cell_val:
        sheet.cell(row = dict["row"], column = dict["col"]).value = dict["value"]
    wb.save(file)

import openpyxl

def delete_row(file,rows):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for row in rows:
        sheet.delete_cols(idx=row)
    path1 = 'final.xlsx'
    wb.save(file)


def delete_column(file,cols):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for col in cols:
        sheet.delete_cols(idx=col)
    path1 = 'final.xlsx'
    wb.save(file)


def merge_tables(files):
    excels = [pd.ExcelFile(name) for name in files]
    frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
    combined = pd.concat(frames)
    combined.to_excel(files[0], header=False, index=False)
    print("merged tables succesfully")
    return files[0]


def merge_cells(file,cells):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    value = ""
    start = cells[0]
    end = cells[1]
    st_r= start["row"]
    st_c= start["col"]
    en_r= end["row"]
    en_c= end["col"]
    for i in range(st_r,en_r+1):
        for j in range(st_c,en_c+1):
            if(sheet.cell(row = i, column = j).value!=None):
                value = value + " " + str(sheet.cell(row = i, column = j).value)
            else:
                value = value + " "
        value =  value + " "
    start = openpyxl.utils.cell.get_column_letter(st_c)
    end = openpyxl.utils.cell.get_column_letter(en_c)
    merge = str(start) + str(st_r) + ":" + str(end) + str(en_r)
    sheet.cell(row = st_r, column = st_c).value = value
    sheet.merge_cells(merge)
    print("merged cells successfully")
    wb.save(file)



def split_cols(file,column,separation_point):
    '''
    file: file to change
    column: column that is split
    separation_point: list of dictionaries of values in the split cells( Format: {"right_cell":"right","left_cell":"left"})'''
    value_r = []
    value_l = []
    for i in separation_point:
        value_r.append(i["right_cell"])
        value_l.append(i["left_cell"])
    insert_column(file,column,value_r)
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for row in range(1,len(value_l)+1):
        sheet.cell(row = row, column = column+1).value = value_l[row-1]
    print("split columns successfully")
    wb.save(file)


def insert_row(file,loc,data):
    '''
    args:
        file: file that needs to be edited
        data: List with values  to insert
        loc: Location to insert the row
    '''
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    sheet.insert_rows(loc)
    sheet._current_row = loc - 1
    sheet.append(data)
    wb.save(file)

def insert_column(file,loc,data):

    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    sheet.insert_cols(loc)
    for row in range(1,len(data)+1):
        sheet.cell(row = row, column = loc).value = data[row-1]
    wb.save(file)



# change cell value
change_cell("/home/vaibhav/Desktop/vivitsa-workspace/table-extraction/aaa/5-0.csv.xlsx",[{"row":1,"col":1,"value":"changed"}])



# split columns
path = "/home/vaibhav/Desktop/vivitsa-workspace/table-extraction/aaa/19-1.csv.xlsx"
split_cols(path,4,[{"right_cell":"","left_cell":""},{"right_cell":"(+) Total","left_cell":"Principal Collected"},{"right_cell":"(-) Total","left_cell":"Losses"},{"right_cell":"(+) Total","left_cell":"Interest Collected"},{"right_cell":"(+) Total Other","left_cell":"Interest Adjust. Collected"},{"right_cell":"","left_cell":""},{"right_cell":"(-) Total","left_cell":"Fees (Withheld)"},{"right_cell":"","left_cell":""},{"right_cell":"(+) Prepayment","left_cell":"Penalty"},{"right_cell":"","left_cell":""},{"right_cell":"Total Available","left_cell":"Funds from Collection"}])

# # merge columns
# for i in range(1,13):
#     merge_cells(path,[{"row":1,"col":i},{"row":3,"col":i}])

# # merge tables
# path0 = "/home/vaibhav/Desktop/vivitsa-workspace/table-extraction/aaa/3-0.csv.xlsx"
# path1 = "/home/vaibhav/Desktop/vivitsa-workspace/table-extraction/aaa/4-0.csv.xlsx"
# path2 = "/home/vaibhav/Desktop/vivitsa-workspace/table-extraction/aaa/5-0.csv.xlsx"
# merge_tables([path0,path1,path2])

# path3 = "/home/vaibhav/Desktop/vivitsa-workspace/table-extraction/aaa/6-0.csv.xlsx"
# insert_column(path3,2,["a","b","c","d","e",'f','g','h','i'])

